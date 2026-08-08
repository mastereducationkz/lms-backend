"""Bluebook group grid: column coverage, cell states and statistics.

The grid must always show every Bluebook test 4-11 and must distinguish three reasons a
cell is empty:

  submitted      - a real score
  not_submitted  - the test WAS assigned to this group; the student did not submit
  not_assigned   - no Bluebook homework exists for that test in this group

Collapsing those into a blank cell hides whether a gap is the student's or the
teacher's, which is the whole point of the view.
"""
import json
from datetime import datetime, timezone

import pytest

from src.schemas.models import (  # noqa: F401  (import-order guard: shim first)
    Assignment,
    AssignmentSubmission,
    Group,
    GroupStudent,
    UserInDB,
)
from src.exams.models import BluebookResult
from src.exams.services import build_bluebook_grid


@pytest.fixture
def db():
    from sqlalchemy import event
    from sqlalchemy.exc import OperationalError
    from sqlalchemy.orm import Session as SASession
    from src.config import engine
    try:
        connection = engine.connect()
    except OperationalError:
        pytest.skip("No database available (requires Postgres); skipping")
    trans = connection.begin()
    session = SASession(bind=connection)
    session.begin_nested()

    @event.listens_for(session, "after_transaction_end")
    def _restart_savepoint(sess, transaction):
        if transaction.nested and not transaction._parent.nested:
            sess.begin_nested()

    try:
        yield session
    finally:
        event.remove(session, "after_transaction_end", _restart_savepoint)
        session.close()
        trans.rollback()
        connection.close()


def _student(db, email, name):
    u = UserInDB(email=email, name=name, hashed_password="x", role="student", is_active=True)
    db.add(u)
    db.flush()
    return u


def _bluebook_assignment(db, group, test_number, due):
    content = {"tasks": [{
        "id": f"task_{test_number}", "task_type": "bluebook_task",
        "title": f"Bluebook Test #{test_number}", "order_index": 0, "points": 10,
        "content": {"test_number": test_number},
    }]}
    a = Assignment(
        title=f"Bluebook #{test_number}", description="", assignment_type="multi_task",
        content=json.dumps(content), max_score=10, group_id=group.id,
        due_date=due, is_active=True,
    )
    db.add(a)
    db.flush()
    return a


def _result(db, student, group, *, assignment=None, test_number=5, verbal=500, math=500,
            source="homework"):
    r = BluebookResult(
        student_id=student.id,
        assignment_id=assignment.id if assignment else None,
        group_id=group.id if assignment else None,
        test_number=test_number,
        verbal_score=verbal, math_score=math, total_score=verbal + math,
        source=source,
    )
    db.add(r)
    db.flush()
    return r


@pytest.fixture
def group_with_students(db):
    g = Group(name="grid Group", is_active=True, is_over=False, program_type="sat")
    db.add(g)
    db.flush()
    a = _student(db, "grid-a@t.io", "AAA Student")
    b = _student(db, "grid-b@t.io", "BBB Student")
    db.add(GroupStudent(group_id=g.id, student_id=a.id))
    db.add(GroupStudent(group_id=g.id, student_id=b.id))
    db.flush()
    return g, a, b


# --------------------------------------------------------------------------------------
# Column coverage
# --------------------------------------------------------------------------------------

def test_all_tests_4_to_11_are_always_columns_even_with_no_data(db, group_with_students):
    """REGRESSION: the grid used to derive columns from existing results, so a group
    with only an Assignment Zero baseline collapsed to a single column."""
    group, _, _ = group_with_students
    grid = build_bluebook_grid(db, group)
    numbers = [c.test_number for c in grid.columns if not c.is_baseline]
    assert numbers == [4, 5, 6, 7, 8, 9, 10, 11]


def test_unassigned_columns_are_flagged_not_assigned(db, group_with_students):
    group, _, _ = group_with_students
    grid = build_bluebook_grid(db, group)
    assert all(c.is_assigned is False for c in grid.columns if not c.is_baseline)


def test_assigning_a_test_marks_only_that_column_assigned(db, group_with_students):
    group, _, _ = group_with_students
    _bluebook_assignment(db, group, 7, datetime(2026, 7, 6, 12, 0))
    grid = build_bluebook_grid(db, group)
    by_number = {c.test_number: c for c in grid.columns if not c.is_baseline}
    assert by_number[7].is_assigned is True
    assert by_number[7].due_date is not None
    assert by_number[4].is_assigned is False


def test_baseline_column_appears_only_when_a_student_has_one(db, group_with_students):
    group, a, _ = group_with_students
    assert not any(c.is_baseline for c in build_bluebook_grid(db, group).columns)

    _result(db, a, group, assignment=None, test_number=5, verbal=330, math=570,
            source="assignment_zero")
    grid = build_bluebook_grid(db, group)
    baseline = [c for c in grid.columns if c.is_baseline]
    assert len(baseline) == 1
    assert baseline[0].test_number == 5


# --------------------------------------------------------------------------------------
# Cell states - the three-way distinction
# --------------------------------------------------------------------------------------

def test_cell_is_not_assigned_when_no_homework_exists(db, group_with_students):
    group, a, _ = group_with_students
    grid = build_bluebook_grid(db, group)
    row = next(r for r in grid.rows if r.student_id == a.id)
    assert row.cells["t4"]["state"] == "not_assigned"
    assert row.cells["t4"]["total_score"] is None


def test_cell_is_not_submitted_when_assigned_but_no_result(db, group_with_students):
    """The distinction that matters: this gap is the student's, not the teacher's."""
    group, a, _ = group_with_students
    _bluebook_assignment(db, group, 6, datetime(2026, 6, 22, 12, 0))
    grid = build_bluebook_grid(db, group)
    row = next(r for r in grid.rows if r.student_id == a.id)
    assert row.cells["t6"]["state"] == "not_submitted"
    assert row.cells["t6"]["total_score"] is None


def test_cell_is_submitted_with_scores_when_a_result_exists(db, group_with_students):
    group, a, _ = group_with_students
    assignment = _bluebook_assignment(db, group, 8, datetime(2026, 7, 20, 12, 0))
    _result(db, a, group, assignment=assignment, test_number=8, verbal=640, math=780)
    grid = build_bluebook_grid(db, group)
    row = next(r for r in grid.rows if r.student_id == a.id)
    cell = row.cells["t8"]
    assert cell["state"] == "submitted"
    assert (cell["verbal_score"], cell["math_score"], cell["total_score"]) == (640, 780, 1420)


def test_one_student_submitting_does_not_mark_the_other_submitted(db, group_with_students):
    group, a, b = group_with_students
    assignment = _bluebook_assignment(db, group, 9, datetime(2026, 8, 3, 12, 0))
    _result(db, a, group, assignment=assignment, test_number=9, verbal=600, math=600)
    grid = build_bluebook_grid(db, group)
    rows = {r.student_id: r for r in grid.rows}
    assert rows[a.id].cells["t9"]["state"] == "submitted"
    assert rows[b.id].cells["t9"]["state"] == "not_submitted"


# --------------------------------------------------------------------------------------
# Statistics
# --------------------------------------------------------------------------------------

def test_unassigned_tests_do_not_count_against_completion(db, group_with_students):
    """Two students, one assigned test, one submission => 50%, not 1/16."""
    group, a, _ = group_with_students
    assignment = _bluebook_assignment(db, group, 4, datetime(2026, 6, 8, 12, 0))
    _result(db, a, group, assignment=assignment, test_number=4, verbal=500, math=500)

    grid = build_bluebook_grid(db, group)
    assert grid.group_stats["tests_assigned"] == 1
    assert grid.group_stats["expected_count"] == 2
    assert grid.group_stats["submitted_count"] == 1
    assert grid.group_stats["completion_rate"] == 0.5


def test_group_stats_report_averages_and_spread(db, group_with_students):
    group, a, b = group_with_students
    assignment = _bluebook_assignment(db, group, 4, datetime(2026, 6, 8, 12, 0))
    _result(db, a, group, assignment=assignment, test_number=4, verbal=400, math=400)  # 800
    _result(db, b, group, assignment=assignment, test_number=4, verbal=600, math=600)  # 1200

    stats = build_bluebook_grid(db, group).group_stats
    assert stats["average_latest_total"] == 1000.0
    assert stats["highest_total"] == 1200
    assert stats["lowest_latest_total"] == 800
    assert stats["students_with_no_results"] == 0


def test_group_stats_count_students_with_no_results(db, group_with_students):
    group, a, _ = group_with_students
    assignment = _bluebook_assignment(db, group, 4, datetime(2026, 6, 8, 12, 0))
    _result(db, a, group, assignment=assignment, test_number=4, verbal=500, math=500)
    assert build_bluebook_grid(db, group).group_stats["students_with_no_results"] == 1


def test_group_stats_are_safe_on_an_empty_group(db):
    """No students and no assignments must not divide by zero."""
    g = Group(name="grid Empty", is_active=True, is_over=False, program_type="sat")
    db.add(g)
    db.flush()
    stats = build_bluebook_grid(db, g).group_stats
    assert stats["student_count"] == 0
    assert stats["completion_rate"] == 0.0
    assert stats["average_latest_total"] is None


def test_student_stats_track_progress_from_the_baseline(db, group_with_students):
    group, a, _ = group_with_students
    _result(db, a, group, assignment=None, test_number=5, verbal=330, math=570,
            source="assignment_zero")                                   # baseline 900
    later = _bluebook_assignment(db, group, 7, datetime(2026, 7, 6, 12, 0))
    _result(db, a, group, assignment=later, test_number=7, verbal=500, math=600)  # 1100

    row = next(r for r in build_bluebook_grid(db, group).rows if r.student_id == a.id)
    assert row.baseline_total == 900
    assert row.latest_total == 1100
    assert row.best_total == 1100
    assert row.improvement_from_baseline == 200
    assert row.submitted_count == 1     # baseline excluded - it is not homework
    assert row.assigned_count == 1


def test_student_average_excludes_missing_cells(db, group_with_students):
    group, a, _ = group_with_students
    a1 = _bluebook_assignment(db, group, 4, datetime(2026, 6, 8, 12, 0))
    a2 = _bluebook_assignment(db, group, 6, datetime(2026, 6, 22, 12, 0))
    _result(db, a, group, assignment=a1, test_number=4, verbal=400, math=400)   # 800
    _result(db, a, group, assignment=a2, test_number=6, verbal=600, math=600)   # 1200

    row = next(r for r in build_bluebook_grid(db, group).rows if r.student_id == a.id)
    assert row.average_total == 1000.0
    assert row.submitted_count == 2
    assert row.assigned_count == 2


def test_trend_compares_against_the_previous_submitted_cell(db, group_with_students):
    """Gaps must not reset or corrupt the trend: 800 -> (gap) -> 1200 is still 'up'."""
    group, a, _ = group_with_students
    a1 = _bluebook_assignment(db, group, 4, datetime(2026, 6, 8, 12, 0))
    _bluebook_assignment(db, group, 6, datetime(2026, 6, 22, 12, 0))  # assigned, skipped
    a3 = _bluebook_assignment(db, group, 8, datetime(2026, 7, 20, 12, 0))
    _result(db, a, group, assignment=a1, test_number=4, verbal=400, math=400)
    _result(db, a, group, assignment=a3, test_number=8, verbal=600, math=600)

    row = next(r for r in build_bluebook_grid(db, group).rows if r.student_id == a.id)
    assert row.cells["t4"]["trend"] is None      # nothing before it
    assert row.cells["t6"]["state"] == "not_submitted"
    assert row.cells["t8"]["trend"] == "up"
    assert row.cells["t8"]["delta"] == 400


def test_students_are_listed_even_with_no_results_at_all(db, group_with_students):
    group, a, b = group_with_students
    grid = build_bluebook_grid(db, group)
    assert {r.student_id for r in grid.rows} == {a.id, b.id}
    assert all(r.latest_total is None for r in grid.rows)


def test_export_writes_not_assigned_and_not_submitted_in_words(db, group_with_students):
    """A blank cell in a spreadsheet is ambiguous; the reason must be spelled out."""
    from io import BytesIO
    from openpyxl import load_workbook
    from src.exams.exports import build_bluebook_grid_workbook

    group, a, _ = group_with_students
    assigned = _bluebook_assignment(db, group, 4, datetime(2026, 6, 8, 12, 0))
    _result(db, a, group, assignment=assigned, test_number=4, verbal=400, math=400)

    grid = build_bluebook_grid(db, group)
    ws = load_workbook(BytesIO(build_bluebook_grid_workbook(grid).getvalue())).active
    text = " ".join(
        str(c.value) for r in ws.iter_rows() for c in r if c.value is not None
    ).lower()

    assert "not assigned" in text     # tests 5-11 were never assigned
    assert "not submitted" in text    # student B was assigned #4 and did not submit
    assert "800" in text              # student A's real score still present
