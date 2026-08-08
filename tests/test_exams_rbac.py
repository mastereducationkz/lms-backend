"""Row-level authorization for the exam-results and Bluebook endpoints.

Covers every surface that can leak a row: list, export, create, update, grid and grid
export. Export is tested explicitly because an export that re-derives scope differently
from the screen is the classic way row-level security is bypassed.

Route functions are called directly with an injected ``current_user``, matching the
established style in this repo (there is no conftest, no TestClient auth, no JWT
minting in permission tests).
"""
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal

import pytest
from fastapi import HTTPException

# The shim must be imported before any domain model module: importing a domain model
# first re-enters the partially-initialized src.models package and raises ImportError.
from src.schemas.models import ExamResult  # noqa: F401  (import-order guard)
from src.exams.routes import (
    create_exam_result,
    export_bluebook_grid,
    export_exam_results,
    get_bluebook_grid,
    list_bluebook_groups,
    list_exam_groups,
    list_exam_results,
    list_sat_official_dates,
    update_exam_result,
)
from src.exams.routes import get_result_proof, update_planned_date, upload_result_proof
from src.exams.schemas import ExamResultCreate, ExamResultUpdate, PlannedDateUpdate
from src.schemas.models import (
    Course,
    CourseGroupAccess,
    CourseHeadTeacher,
    Group,
    GroupStudent,
    UserInDB,
)


@pytest.fixture
def db():
    from sqlalchemy import event
    from sqlalchemy.exc import OperationalError
    from sqlalchemy.orm import Session as SASession
    from src.config import engine
    try:
        connection = engine.connect()
    except OperationalError:
        pytest.skip("No database available (requires Postgres); skipping")
    trans = connection.begin()
    session = SASession(bind=connection)
    session.begin_nested()

    @event.listens_for(session, "after_transaction_end")
    def _restart_savepoint(sess, transaction):
        if transaction.nested and not transaction._parent.nested:
            sess.begin_nested()

    try:
        yield session
    finally:
        event.remove(session, "after_transaction_end", _restart_savepoint)
        session.close()
        trans.rollback()
        connection.close()


def _user(db, role, email, name=None):
    u = UserInDB(email=email, name=name or f"{role} {email}", hashed_password="x",
                 role=role, is_active=True)
    db.add(u)
    db.flush()
    return u


def _group(db, name, *, teacher_id=None, curator_id=None):
    g = Group(name=name, teacher_id=teacher_id, curator_id=curator_id,
              is_active=True, is_over=False, program_type="sat")
    db.add(g)
    db.flush()
    return g


def _enrol(db, group, student):
    db.add(GroupStudent(group_id=group.id, student_id=student.id))
    db.flush()


def _result(db, student, *, test_date=None, total=1450):
    r = ExamResult(
        student_id=student.id, exam_type="sat",
        test_date=test_date or date(2026, 6, 6),
        total_score=Decimal(total), verbal_score=700, math_score=750,
        status="reported", source="staff",
        recorded_at=datetime.now(timezone.utc), is_superseded=False,
    )
    db.add(r)
    db.flush()
    return r


@pytest.fixture
def world(db):
    """Two disjoint groups with their own staff and students."""
    t_a = _user(db, "teacher", "rb-ta@t.io")
    t_b = _user(db, "teacher", "rb-tb@t.io")
    c_a = _user(db, "curator", "rb-ca@t.io")
    c_b = _user(db, "curator", "rb-cb@t.io")

    g_a = _group(db, "rb Group A", teacher_id=t_a.id, curator_id=c_a.id)
    g_b = _group(db, "rb Group B", teacher_id=t_b.id, curator_id=c_b.id)

    s_a = _user(db, "student", "rb-sa@t.io", name="rb Student A")
    s_b = _user(db, "student", "rb-sb@t.io", name="rb Student B")
    _enrol(db, g_a, s_a)
    _enrol(db, g_b, s_b)

    _result(db, s_a, total=1450)
    _result(db, s_b, total=1300)

    return dict(t_a=t_a, t_b=t_b, c_a=c_a, c_b=c_b,
                g_a=g_a, g_b=g_b, s_a=s_a, s_b=s_b)


def _ids(rows):
    return {r.student.student_id for r in rows}


def _list(user, db, **kw):
    params = dict(exam_type="sat", group_id=None, date_field="planned",
                  date_from=None, date_to=None, exact_date=None,
                  status=None, search=None, limit=200, offset=0,
                  current_user=user, db=db)
    params.update(kw)
    return list_exam_results(**params)


# --------------------------------------------------------------------------------------
# Role gate
# --------------------------------------------------------------------------------------

@pytest.mark.parametrize("role", ["student", "parent"])
def test_non_staff_roles_are_denied_the_results_list(db, world, role):
    u = _user(db, role, f"rb-deny-{role}@t.io")
    with pytest.raises(HTTPException) as exc:
        _list(u, db)
    assert exc.value.status_code == 403


def test_unknown_role_is_denied(db, world):
    u = _user(db, "marketing", "rb-deny-mkt@t.io")
    with pytest.raises(HTTPException) as exc:
        _list(u, db)
    assert exc.value.status_code == 403


def test_teacher_may_read_but_not_write(db, world):
    """Teachers see results; recording an official score is curator-owned."""
    _list(world["t_a"], db)  # no raise

    with pytest.raises(HTTPException) as exc:
        create_exam_result(
            ExamResultCreate(student_id=world["s_a"].id, exam_type="sat",
                             test_date=date(2026, 5, 2), verbal_score=600, math_score=600),
            current_user=world["t_a"], db=db,
        )
    assert exc.value.status_code == 403


# --------------------------------------------------------------------------------------
# Row scope - list
# --------------------------------------------------------------------------------------

def test_teacher_sees_only_own_group_students(db, world):
    rows = _list(world["t_a"], db)
    assert world["s_a"].id in _ids(rows)
    assert world["s_b"].id not in _ids(rows)


def test_curator_sees_only_own_group_students(db, world):
    rows = _list(world["c_a"], db)
    assert world["s_a"].id in _ids(rows)
    assert world["s_b"].id not in _ids(rows)


def test_admin_sees_both_groups(db, world):
    admin = _user(db, "admin", "rb-admin@t.io")
    ids = _ids(_list(admin, db))
    assert world["s_a"].id in ids
    assert world["s_b"].id in ids


def test_head_curator_sees_both_groups(db, world):
    hc = _user(db, "head_curator", "rb-hc@t.io")
    ids = _ids(_list(hc, db))
    assert world["s_a"].id in ids
    assert world["s_b"].id in ids


def test_staff_with_no_groups_sees_nothing_not_everything(db, world):
    """The classic row-scope failure: an empty scope list is falsy, and code that
    checks truthiness instead of the UNRESTRICTED sentinel silently degrades to
    'no filter' - handing a brand-new curator every student in the LMS."""
    lonely = _user(db, "curator", "rb-lonely@t.io")
    assert _list(lonely, db) == []


def test_staff_with_no_groups_cannot_create_a_result(db, world):
    lonely = _user(db, "curator", "rb-lonely2@t.io")
    with pytest.raises(HTTPException) as exc:
        create_exam_result(
            ExamResultCreate(student_id=world["s_a"].id, exam_type="sat",
                             test_date=date(2026, 5, 2), verbal_score=600, math_score=600),
            current_user=lonely, db=db,
        )
    assert exc.value.status_code == 403


def test_staff_with_no_groups_exports_an_empty_workbook_not_everything(db, world):
    lonely = _user(db, "curator", "rb-lonely3@t.io")
    resp = export_exam_results(
        exam_type="sat", group_id=None, date_field="planned",
        date_from=None, date_to=None, exact_date=None, status=None, search=None,
        current_user=lonely, db=db,
    )
    assert resp.body[:2] == b"PK"
    # A workbook with only the header row - no student data reached it.
    from io import BytesIO
    from openpyxl import load_workbook
    ws = load_workbook(BytesIO(resp.body)).active
    assert ws.max_row == 1


def test_requesting_a_foreign_group_is_403_not_an_empty_list(db, world):
    """A silent empty result would hide the authorization failure from the operator."""
    with pytest.raises(HTTPException) as exc:
        _list(world["t_a"], db, group_id=world["g_b"].id)
    assert exc.value.status_code == 403


def test_teacher_may_filter_to_their_own_group(db, world):
    rows = _list(world["t_a"], db, group_id=world["g_a"].id)
    assert _ids(rows) == {world["s_a"].id}


# --------------------------------------------------------------------------------------
# Row scope - export must match the screen exactly
# --------------------------------------------------------------------------------------

def test_export_cannot_reach_rows_the_grid_hides(db, world):
    """The export re-derives scope from the user, so a foreign group_id is refused."""
    with pytest.raises(HTTPException) as exc:
        export_exam_results(
            exam_type="sat", group_id=world["g_b"].id, date_field="planned",
            date_from=None, date_to=None, exact_date=None, status=None, search=None,
            current_user=world["t_a"], db=db,
        )
    assert exc.value.status_code == 403


def test_export_is_denied_for_non_staff(db, world):
    student = _user(db, "student", "rb-exp-stu@t.io")
    with pytest.raises(HTTPException) as exc:
        export_exam_results(
            exam_type="sat", group_id=None, date_field="planned",
            date_from=None, date_to=None, exact_date=None, status=None, search=None,
            current_user=student, db=db,
        )
    assert exc.value.status_code == 403


def test_export_produces_a_workbook_for_an_authorized_user(db, world):
    resp = export_exam_results(
        exam_type="sat", group_id=None, date_field="planned",
        date_from=None, date_to=None, exact_date=None, status=None, search=None,
        current_user=world["c_a"], db=db,
    )
    # XLSX files are ZIP archives; check the magic bytes rather than trusting status.
    assert resp.body[:2] == b"PK"
    assert "attachment" in resp.headers["Content-Disposition"]


# --------------------------------------------------------------------------------------
# Row scope - create / update
# --------------------------------------------------------------------------------------

def test_curator_cannot_create_a_result_for_a_foreign_student(db, world):
    with pytest.raises(HTTPException) as exc:
        create_exam_result(
            ExamResultCreate(student_id=world["s_b"].id, exam_type="sat",
                             test_date=date(2026, 5, 2), verbal_score=600, math_score=600),
            current_user=world["c_a"], db=db,
        )
    assert exc.value.status_code == 403


def test_curator_can_create_for_own_student(db, world):
    out = create_exam_result(
        ExamResultCreate(student_id=world["s_a"].id, exam_type="sat",
                         test_date=date(2026, 5, 2), verbal_score=600, math_score=610),
        current_user=world["c_a"], db=db,
    )
    assert out.total_score == Decimal("1210")
    assert out.status == "reported"


def test_duplicate_attempt_for_same_date_is_rejected(db, world):
    with pytest.raises(HTTPException) as exc:
        create_exam_result(
            ExamResultCreate(student_id=world["s_a"].id, exam_type="sat",
                             test_date=date(2026, 6, 6), verbal_score=600, math_score=600),
            current_user=world["c_a"], db=db,
        )
    assert exc.value.status_code == 409


def test_multiple_attempts_on_different_dates_are_allowed(db, world):
    """The whole point of the new table: a retake must not overwrite the first score."""
    create_exam_result(
        ExamResultCreate(student_id=world["s_a"].id, exam_type="sat",
                         test_date=date(2026, 5, 2), verbal_score=600, math_score=610),
        current_user=world["c_a"], db=db,
    )
    kept = db.query(ExamResult).filter(ExamResult.student_id == world["s_a"].id).count()
    assert kept == 2


def test_curator_cannot_update_a_foreign_students_result(db, world):
    foreign = db.query(ExamResult).filter(ExamResult.student_id == world["s_b"].id).first()
    with pytest.raises(HTTPException) as exc:
        update_exam_result(foreign.id, ExamResultUpdate(status="verified"),
                           current_user=world["c_a"], db=db)
    assert exc.value.status_code == 403


def test_verifying_records_the_actor(db, world):
    own = db.query(ExamResult).filter(ExamResult.student_id == world["s_a"].id).first()
    out = update_exam_result(own.id, ExamResultUpdate(status="verified"),
                             current_user=world["c_a"], db=db)
    assert out.status == "verified"
    refreshed = db.query(ExamResult).filter(ExamResult.id == own.id).first()
    assert refreshed.verified_by == world["c_a"].id
    assert refreshed.verified_at is not None


# --------------------------------------------------------------------------------------
# Bluebook grid scope
# --------------------------------------------------------------------------------------

# --------------------------------------------------------------------------------------
# Bluebook group selector - the list that populates the grid's dropdown
# --------------------------------------------------------------------------------------

def _group_ids(rows):
    return {r["id"] for r in rows}


def test_admin_sees_sat_groups_in_the_selector(db, world):
    """REGRESSION: the page originally used GET /users/groups/me, which has no admin
    branch and returns [], so an admin saw 'You have no SAT groups'."""
    admin = _user(db, "admin", "bg-admin@t.io")
    rows = list_bluebook_groups(search=None, current_user=admin, db=db)
    assert world["g_a"].id in _group_ids(rows)
    assert world["g_b"].id in _group_ids(rows)


def test_head_curator_sees_sat_groups_in_the_selector(db, world):
    """Same regression: /users/groups/me has no head_curator branch either."""
    hc = _user(db, "head_curator", "bg-hc@t.io")
    assert world["g_a"].id in _group_ids(list_bluebook_groups(search=None, current_user=hc, db=db))


def test_head_teacher_sees_only_course_linked_groups_in_the_selector(db, world):
    """Same regression for head_teacher, and it must still respect course-link scope."""
    ht = _user(db, "head_teacher", "bg-ht@t.io")
    assert list_bluebook_groups(search=None, current_user=ht, db=db) == []

    course = Course(title="bg Course", description="d", teacher_id=ht.id)
    db.add(course)
    db.flush()
    db.add(CourseHeadTeacher(course_id=course.id, head_teacher_id=ht.id))
    db.add(CourseGroupAccess(course_id=course.id, group_id=world["g_a"].id,
                             granted_by=ht.id, is_active=True))
    db.flush()

    ids = _group_ids(list_bluebook_groups(search=None, current_user=ht, db=db))
    assert world["g_a"].id in ids
    assert world["g_b"].id not in ids


def test_teacher_sees_only_own_groups_in_the_selector(db, world):
    ids = _group_ids(list_bluebook_groups(search=None, current_user=world["t_a"], db=db))
    assert ids == {world["g_a"].id}


def test_nuet_groups_are_excluded_because_bluebook_is_sat_only(db, world):
    """Bluebook is a College Board SAT product; NUET students do not sit it."""
    admin = _user(db, "admin", "bg-admin2@t.io")
    nuet = Group(name="bg NUET squad", is_active=True, is_over=False, program_type="nuet")
    db.add(nuet)
    db.flush()
    assert nuet.id not in _group_ids(list_bluebook_groups(search=None, current_user=admin, db=db))


def test_general_english_group_named_saturday_is_not_matched_as_sat(db, world):
    """The name fallback uses a word boundary, so 'Saturday' must not read as 'SAT'."""
    admin = _user(db, "admin", "bg-admin3@t.io")
    sat_urday = Group(name="bg Saturday morning club", is_active=True, is_over=False,
                      program_type="general_english")
    db.add(sat_urday)
    db.flush()
    assert sat_urday.id not in _group_ids(list_bluebook_groups(search=None, current_user=admin, db=db))


def test_legacy_group_without_program_type_is_matched_by_name(db, world):
    admin = _user(db, "admin", "bg-admin4@t.io")
    legacy = Group(name="bg SAT August intake", is_active=True, is_over=False,
                   program_type="general_english")
    db.add(legacy)
    db.flush()
    assert legacy.id in _group_ids(list_bluebook_groups(search=None, current_user=admin, db=db))


def test_selector_is_denied_for_students(db, world):
    student = _user(db, "student", "bg-stu@t.io")
    with pytest.raises(HTTPException) as exc:
        list_bluebook_groups(search=None, current_user=student, db=db)
    assert exc.value.status_code == 403


def test_teacher_cannot_open_another_groups_bluebook_grid(db, world):
    with pytest.raises(HTTPException) as exc:
        get_bluebook_grid(world["g_b"].id, cohort_date=None,
                          current_user=world["t_a"], db=db)
    assert exc.value.status_code == 403


def test_teacher_can_open_own_bluebook_grid(db, world):
    grid = get_bluebook_grid(world["g_a"].id, cohort_date=None,
                             current_user=world["t_a"], db=db)
    assert grid.group_id == world["g_a"].id
    assert [r.student_id for r in grid.rows] == [world["s_a"].id]


def test_bluebook_grid_export_enforces_the_same_scope(db, world):
    with pytest.raises(HTTPException) as exc:
        export_bluebook_grid(world["g_b"].id, cohort_date=None,
                             current_user=world["t_a"], db=db)
    assert exc.value.status_code == 403


def test_head_teacher_without_course_link_cannot_see_a_sat_group(db, world):
    """Course-link scope, not program_type scope: managing nothing means seeing
    nothing, even for a group of a program they nominally head."""
    ht = _user(db, "head_teacher", "rb-ht@t.io")
    with pytest.raises(HTTPException) as exc:
        get_bluebook_grid(world["g_a"].id, cohort_date=None, current_user=ht, db=db)
    assert exc.value.status_code == 403


def test_head_teacher_with_course_link_can_see_the_linked_group(db, world):
    ht = _user(db, "head_teacher", "rb-ht2@t.io")
    course = Course(title="rb Course", description="d", teacher_id=ht.id)
    db.add(course)
    db.flush()
    db.add(CourseHeadTeacher(course_id=course.id, head_teacher_id=ht.id))
    db.add(CourseGroupAccess(course_id=course.id, group_id=world["g_a"].id,
                             granted_by=ht.id, is_active=True))
    db.flush()

    grid = get_bluebook_grid(world["g_a"].id, cohort_date=None, current_user=ht, db=db)
    assert grid.group_id == world["g_a"].id


# --------------------------------------------------------------------------------------
# Payload minimization
# --------------------------------------------------------------------------------------

def test_result_rows_never_expose_college_board_credentials(db, world):
    """The existing AssignmentZeroSubmissionSchema declares these fields and is
    inherited by CuratorUpcomingExamRow. Nothing in this domain may reuse it."""
    rows = _list(world["c_a"], db)
    assert rows
    for row in rows:
        blob = row.model_dump_json()
        assert "college_board" not in blob
        assert "password" not in blob.lower()


def test_result_rows_do_not_leak_the_proof_storage_key(db, world):
    """Proof of an official exam is a score report - a list payload exposes only
    whether one exists."""
    own = db.query(ExamResult).filter(ExamResult.student_id == world["s_a"].id).first()
    own.proof_url = "exam_proof/secret-key-123.pdf"
    db.flush()

    rows = _list(world["c_a"], db)
    for row in rows:
        blob = row.model_dump_json()
        assert "secret-key-123" not in blob


# --------------------------------------------------------------------------------------
# SAT dates endpoint
# --------------------------------------------------------------------------------------

def test_sat_dates_excludes_anticipated_by_default(db, world):
    payload = list_sat_official_dates(include_anticipated=False, include_past=True,
                                      current_user=world["t_a"], db=db)
    assert all(d["status"] == "confirmed" for d in payload["dates"])


def test_sat_dates_labels_anticipated_when_requested(db, world):
    payload = list_sat_official_dates(include_anticipated=True, include_past=True,
                                      current_user=world["t_a"], db=db)
    statuses = {d["status"] for d in payload["dates"]}
    assert "anticipated" in statuses
    anticipated = [d for d in payload["dates"] if d["status"] == "anticipated"]
    assert {d["test_date"] for d in anticipated} >= {"2027-08-28", "2028-06-03"}


def test_sat_dates_carry_provenance(db, world):
    payload = list_sat_official_dates(include_anticipated=False, include_past=True,
                                      current_user=world["t_a"], db=db)
    assert payload["source_url"].startswith("https://satsuite.collegeboard.org")
    assert payload["verified_at"]


def test_sat_dates_include_registration_deadlines(db, world):
    payload = list_sat_official_dates(include_anticipated=False, include_past=True,
                                      current_user=world["t_a"], db=db)
    aug = next(d for d in payload["dates"] if d["test_date"] == "2026-08-22")
    assert aug["registration_deadline"] == "2026-08-07"
    assert aug["change_deadline"] == "2026-08-11"


def test_group_search_matches_group_name(db, world):
    admin = _user(db, "admin", "bg-search1@t.io")
    rows = list_bluebook_groups(search="Group A", current_user=admin, db=db)
    ids = _group_ids(rows)
    assert world["g_a"].id in ids
    assert world["g_b"].id not in ids


def test_group_search_matches_teacher_name(db, world):
    """Searching by teacher must work, not just by group name."""
    admin = _user(db, "admin", "bg-search2@t.io")
    teacher = db.query(UserInDB).filter(UserInDB.id == world["g_a"].teacher_id).one()
    teacher.name = "Azamat Abduraimov"
    db.flush()

    ids = _group_ids(list_bluebook_groups(search="Abduraimov", current_user=admin, db=db))
    assert world["g_a"].id in ids
    assert world["g_b"].id not in ids


def test_group_search_is_case_insensitive(db, world):
    admin = _user(db, "admin", "bg-search3@t.io")
    assert world["g_a"].id in _group_ids(
        list_bluebook_groups(search="gROUP a", current_user=admin, db=db))


def test_group_search_still_respects_row_scope(db, world):
    """Search must never widen scope - a teacher searching for another group gets none."""
    rows = list_bluebook_groups(search="Group B", current_user=world["t_a"], db=db)
    assert rows == []


# --------------------------------------------------------------------------------------
# /exams/groups - the exam-results group filter (all programs, not just SAT)
# --------------------------------------------------------------------------------------

def test_exam_groups_include_non_sat_programs(db, world):
    """Unlike the Bluebook picker, this one must not be SAT-only - IELTS results are
    filtered by group too."""
    admin = _user(db, "admin", "eg-admin@t.io")
    ielts = Group(name="eg IELTS squad", is_active=True, is_over=False, program_type="ielts")
    db.add(ielts)
    db.flush()
    ids = {r["id"] for r in list_exam_groups(program=None, search=None,
                                             current_user=admin, db=db)}
    assert ielts.id in ids
    assert world["g_a"].id in ids


def test_exam_groups_can_be_narrowed_to_one_program(db, world):
    admin = _user(db, "admin", "eg-admin2@t.io")
    ielts = Group(name="eg IELTS only", is_active=True, is_over=False, program_type="ielts")
    db.add(ielts)
    db.flush()
    ids = {r["id"] for r in list_exam_groups(program="ielts", search=None,
                                             current_user=admin, db=db)}
    assert ielts.id in ids
    assert world["g_a"].id not in ids


def test_exam_groups_respect_row_scope(db, world):
    ids = {r["id"] for r in list_exam_groups(program=None, search=None,
                                             current_user=world["t_a"], db=db)}
    assert ids == {world["g_a"].id}


def test_exam_groups_are_denied_to_students(db, world):
    student = _user(db, "student", "eg-stu@t.io")
    with pytest.raises(HTTPException) as exc:
        list_exam_groups(program=None, search=None, current_user=student, db=db)
    assert exc.value.status_code == 403


# --------------------------------------------------------------------------------------
# Proof upload / retrieval
# --------------------------------------------------------------------------------------

def _png_bytes():
    return b"\x89PNG\r\n\x1a\n" + b"\x00" * 64


class _FakeUpload:
    """Minimal UploadFile stand-in: the handler only calls .read()."""
    def __init__(self, data): self._data = data
    async def read(self): return self._data


def _run(coro):
    """Drive an async handler from a sync test.

    asyncio.run(), not get_event_loop(): another test in the suite closes the ambient
    loop, so get_event_loop() raises "no current event loop" when the full suite runs
    even though these tests pass in isolation.
    """
    import asyncio
    return asyncio.run(coro)


def test_proof_upload_is_denied_to_readers(db, world):
    own = db.query(ExamResult).filter(ExamResult.student_id == world["s_a"].id).first()
    with pytest.raises(HTTPException) as exc:
        _run(upload_result_proof(own.id, file=_FakeUpload(_png_bytes()),
                                 current_user=world["t_a"], db=db))
    assert exc.value.status_code == 403


def test_proof_upload_rejects_a_disguised_executable(db, world):
    """Content-Type is client-supplied; the magic bytes decide."""
    own = db.query(ExamResult).filter(ExamResult.student_id == world["s_a"].id).first()
    with pytest.raises(HTTPException) as exc:
        _run(upload_result_proof(own.id, file=_FakeUpload(b"MZ\x90\x00 not an image"),
                                 current_user=world["c_a"], db=db))
    assert exc.value.status_code == 400


def test_proof_upload_rejects_empty_file(db, world):
    own = db.query(ExamResult).filter(ExamResult.student_id == world["s_a"].id).first()
    with pytest.raises(HTTPException) as exc:
        _run(upload_result_proof(own.id, file=_FakeUpload(b""),
                                 current_user=world["c_a"], db=db))
    assert exc.value.status_code == 400


def test_proof_upload_rejects_oversize(db, world):
    own = db.query(ExamResult).filter(ExamResult.student_id == world["s_a"].id).first()
    huge = b"\x89PNG\r\n\x1a\n" + b"\x00" * (11 * 1024 * 1024)
    with pytest.raises(HTTPException) as exc:
        _run(upload_result_proof(own.id, file=_FakeUpload(huge),
                                 current_user=world["c_a"], db=db))
    assert exc.value.status_code == 400


def test_curator_cannot_attach_proof_to_a_foreign_students_result(db, world):
    foreign = db.query(ExamResult).filter(ExamResult.student_id == world["s_b"].id).first()
    with pytest.raises(HTTPException) as exc:
        _run(upload_result_proof(foreign.id, file=_FakeUpload(_png_bytes()),
                                 current_user=world["c_a"], db=db))
    assert exc.value.status_code == 403


def test_proof_lands_under_the_private_prefix(db, world):
    """exam_proof must be private; assignment_zero screenshots are public-class and
    protected only by an unguessable filename."""
    from src.services.storage_service import is_public
    own = db.query(ExamResult).filter(ExamResult.student_id == world["s_a"].id).first()
    out = _run(upload_result_proof(own.id, file=_FakeUpload(_png_bytes()),
                                   current_user=world["c_a"], db=db))
    assert out.has_proof is True
    refreshed = db.query(ExamResult).filter(ExamResult.id == own.id).first()
    assert "exam_proof/" in refreshed.proof_url
    assert is_public(refreshed.proof_url) is False


def test_proof_retrieval_denied_across_group_boundary(db, world):
    foreign = db.query(ExamResult).filter(ExamResult.student_id == world["s_b"].id).first()
    foreign.proof_url = "/uploads/exam_proof/x.png"
    db.flush()
    with pytest.raises(HTTPException) as exc:
        get_result_proof(foreign.id, current_user=world["t_a"], db=db)
    assert exc.value.status_code == 403


def test_proof_retrieval_404s_when_nothing_uploaded(db, world):
    own = db.query(ExamResult).filter(ExamResult.student_id == world["s_a"].id).first()
    own.proof_url = None
    db.flush()
    with pytest.raises(HTTPException) as exc:
        get_result_proof(own.id, current_user=world["c_a"], db=db)
    assert exc.value.status_code == 404


# --------------------------------------------------------------------------------------
# Planned date + Assignment Zero mirroring
# --------------------------------------------------------------------------------------

def _az(db, student, **kw):
    from src.schemas.models import AssignmentZeroSubmission
    row = AssignmentZeroSubmission(
        user_id=student.id, full_name=student.name, phone_number="", parent_phone_number="",
        telegram_id="", email=student.email, college_board_email="", college_board_password="",
        birthday_date=date(2008, 1, 1), city="Almaty", school_type="NIS", group_name="g",
        sat_target_date="October", recent_practice_test_score="0",
        bluebook_practice_test_5_score="0", **kw,
    )
    db.add(row)
    db.flush()
    return row


def test_recording_a_result_mirrors_into_assignment_zero(db, world):
    """The curator task scheduler closes its task by reading sat_result_score, so a
    result written only to exam_results would leave the task open forever."""
    az = _az(db, world["s_a"])
    # Later than the 2026-06-06 attempt the fixture seeds, so this is the newest.
    create_exam_result(
        ExamResultCreate(student_id=world["s_a"].id, exam_type="sat",
                         test_date=date(2026, 8, 1), verbal_score=600, math_score=610),
        current_user=world["c_a"], db=db,
    )
    db.refresh(az)
    assert az.sat_result_score == "1210"
    assert az.sat_result_test_date == date(2026, 8, 1)


def test_mirroring_follows_the_newest_attempt(db, world):
    az = _az(db, world["s_a"])
    create_exam_result(
        ExamResultCreate(student_id=world["s_a"].id, exam_type="sat",
                         test_date=date(2026, 3, 14), verbal_score=400, math_score=400),
        current_user=world["c_a"], db=db)
    create_exam_result(
        ExamResultCreate(student_id=world["s_a"].id, exam_type="sat",
                         test_date=date(2026, 8, 1), verbal_score=700, math_score=700),
        current_user=world["c_a"], db=db)
    db.refresh(az)
    assert az.sat_result_score == "1400"   # the latest sitting, not an earlier one


def test_rescheduling_does_not_destroy_an_existing_result(db, world):
    """The legacy Assignment Zero endpoint nulls the result on reschedule. With attempt
    history that is data loss - a past sitting really happened."""
    az = _az(db, world["s_a"], sat_result_score="1450",
             sat_result_test_date=date(2026, 6, 6))
    update_planned_date(
        PlannedDateUpdate(student_id=world["s_a"].id, exam_type="sat",
                          planned_test_date=date(2026, 10, 3)),
        current_user=world["c_a"], db=db,
    )
    db.refresh(az)
    assert az.sat_planned_test_date == date(2026, 10, 3)
    assert az.sat_result_score == "1450"          # preserved
    assert az.sat_result_test_date == date(2026, 6, 6)


def test_rescheduling_is_denied_for_a_foreign_student(db, world):
    _az(db, world["s_b"])
    with pytest.raises(HTTPException) as exc:
        update_planned_date(
            PlannedDateUpdate(student_id=world["s_b"].id, exam_type="sat",
                              planned_test_date=date(2026, 10, 3)),
            current_user=world["c_a"], db=db)
    assert exc.value.status_code == 403


def test_rescheduling_is_denied_to_readers(db, world):
    _az(db, world["s_a"])
    with pytest.raises(HTTPException) as exc:
        update_planned_date(
            PlannedDateUpdate(student_id=world["s_a"].id, exam_type="sat",
                              planned_test_date=date(2026, 10, 3)),
            current_user=world["t_a"], db=db)
    assert exc.value.status_code == 403


# --------------------------------------------------------------------------------------
# Attempts + triage on the row
# --------------------------------------------------------------------------------------

def test_rows_carry_full_attempt_history_newest_first(db, world):
    create_exam_result(
        ExamResultCreate(student_id=world["s_a"].id, exam_type="sat",
                         test_date=date(2026, 3, 14), verbal_score=400, math_score=400),
        current_user=world["c_a"], db=db)
    row = next(r for r in _list(world["c_a"], db)
               if r.student.student_id == world["s_a"].id)
    assert len(row.attempts) == 2
    assert row.attempts[0].test_date > row.attempts[1].test_date
    assert row.result.test_date == row.attempts[0].test_date


def test_triage_marks_a_student_with_a_result_completed(db, world):
    row = next(r for r in _list(world["c_a"], db)
               if r.student.student_id == world["s_a"].id)
    assert row.triage_status == "completed"


def test_triage_marks_a_student_without_a_planned_date_unscheduled(db, world):
    db.query(ExamResult).filter(ExamResult.student_id == world["s_a"].id).delete()
    db.flush()
    row = next(r for r in _list(world["c_a"], db)
               if r.student.student_id == world["s_a"].id)
    assert row.triage_status == "unscheduled"
    assert row.ask_result_on is None


def test_ask_date_is_thirteen_days_after_the_planned_date(db, world):
    db.query(ExamResult).filter(ExamResult.student_id == world["s_a"].id).delete()
    _az(db, world["s_a"], sat_planned_test_date=date(2026, 6, 6))
    db.flush()
    row = next(r for r in _list(world["c_a"], db)
               if r.student.student_id == world["s_a"].id)
    assert row.ask_result_on == date(2026, 6, 19)
    assert row.triage_status == "overdue"   # 2026-06-19 is in the past
